"""
CHENU Excel to PostgreSQL Migration Script
Automated migration from Excel database design to PostgreSQL
"""

import pandas as pd
import psycopg2
from psycopg2.extras import execute_values
from openpyxl import load_workbook
import json
from datetime import datetime
from typing import Dict, List, Any

class ExcelToPostgreSQLMigration:
    """
    Migrate CHENU database from Excel design to PostgreSQL
    """
    
    def __init__(self, excel_file: str, connection_string: str):
        """
        Initialize migration
        
        Args:
            excel_file: Path to CHENU_DATABASE_DESIGN.xlsx
            connection_string: PostgreSQL connection string
        """
        self.excel_file = excel_file
        self.connection_string = connection_string
        self.conn = None
        self.wb = None
    
    def connect(self):
        """Connect to PostgreSQL"""
        try:
            self.conn = psycopg2.connect(self.connection_string)
            print("✅ Connected to PostgreSQL")
        except Exception as e:
            print(f"❌ Connection failed: {e}")
            raise
    
    def load_excel(self):
        """Load Excel workbook"""
        try:
            self.wb = load_workbook(self.excel_file)
            print(f"✅ Loaded Excel file: {self.excel_file}")
        except Exception as e:
            print(f"❌ Failed to load Excel: {e}")
            raise
    
    def extract_table_definition(self, sheet_name: str) -> Dict[str, Any]:
        """
        Extract table definition from Excel sheet
        
        Returns:
            {
                'table_name': str,
                'columns': [{'name': str, 'type': str, 'key': str, ...}],
                'sample_data': [[...], [...]]
            }
        """
        ws = self.wb[sheet_name]
        
        # Extract column definitions (rows 2 onwards until "SAMPLE DATA")
        columns = []
        sample_data = []
        in_sample_section = False
        sample_header_row = None
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Check if we've reached sample data section
            if row[0] and "SAMPLE DATA" in str(row[0]):
                in_sample_section = True
                continue
            
            # If in sample section and row is not empty
            if in_sample_section:
                # First non-empty row in sample section is header
                if sample_header_row is None and any(row):
                    sample_header_row = row
                    continue
                
                # Subsequent rows are sample data
                if any(row):
                    sample_data.append(row)
            
            # Otherwise, it's a column definition
            elif row[0]:  # If field name exists
                column = {
                    'name': row[0],
                    'type': row[1],
                    'length': row[2],
                    'key': row[3],
                    'required': row[4] == 'Yes',
                    'default': row[5],
                    'description': row[6],
                    'example': row[7]
                }
                columns.append(column)
        
        # Extract table name from sheet name (e.g., "1_AGENTS" -> "agents")
        table_name = sheet_name.split('_', 1)[1].lower() if '_' in sheet_name else sheet_name.lower()
        
        return {
            'table_name': table_name,
            'columns': columns,
            'sample_data': sample_data
        }
    
    def generate_create_table_sql(self, table_def: Dict[str, Any]) -> str:
        """
        Generate CREATE TABLE SQL from table definition
        """
        table_name = table_def['table_name']
        columns_sql = []
        primary_keys = []
        foreign_keys = []
        
        for col in table_def['columns']:
            col_name = col['name']
            col_type = col['type']
            
            # Map Excel types to PostgreSQL types
            if col_type == 'VARCHAR':
                pg_type = f"VARCHAR({col['length']})" if col['length'] else "VARCHAR(255)"
            elif col_type == 'TEXT':
                pg_type = "TEXT"
            elif col_type == 'INTEGER':
                pg_type = "INTEGER"
            elif col_type == 'BOOLEAN':
                pg_type = "BOOLEAN"
            elif col_type == 'DECIMAL':
                pg_type = f"DECIMAL{col['length']}" if col['length'] else "DECIMAL(10,2)"
            elif col_type == 'TIMESTAMP':
                pg_type = "TIMESTAMP"
            elif col_type == 'JSON' or col_type == 'JSONB':
                pg_type = "JSONB"
            else:
                pg_type = "TEXT"  # Default
            
            # Build column definition
            col_def = f"{col_name} {pg_type}"
            
            # Add NOT NULL if required
            if col['required']:
                col_def += " NOT NULL"
            
            # Add DEFAULT if specified
            if col['default']:
                if col['default'] == 'NOW()':
                    col_def += " DEFAULT NOW()"
                elif col['default'] == 'FALSE':
                    col_def += " DEFAULT FALSE"
                elif col['default'] == 'TRUE':
                    col_def += " DEFAULT TRUE"
                else:
                    col_def += f" DEFAULT '{col['default']}'"
            
            columns_sql.append(col_def)
            
            # Track primary and foreign keys
            if col['key'] == 'PK':
                primary_keys.append(col_name)
            elif col['key'] == 'FK':
                foreign_keys.append(col_name)
        
        # Add PRIMARY KEY constraint
        if primary_keys:
            columns_sql.append(f"PRIMARY KEY ({', '.join(primary_keys)})")
        
        # Combine into CREATE TABLE statement
        sql = f"CREATE TABLE IF NOT EXISTS {table_name} (\n"
        sql += ",\n".join(f"  {col}" for col in columns_sql)
        sql += "\n);"
        
        return sql
    
    def generate_insert_sample_data_sql(self, table_def: Dict[str, Any]) -> List[str]:
        """
        Generate INSERT statements for sample data
        """
        table_name = table_def['table_name']
        columns = table_def['columns']
        sample_data = table_def['sample_data']
        
        if not sample_data:
            return []
        
        insert_statements = []
        column_names = [col['name'] for col in columns]
        
        for row in sample_data:
            # Clean and format values
            values = []
            for idx, val in enumerate(row):
                if val is None or val == '':
                    values.append('NULL')
                elif isinstance(val, bool):
                    values.append('TRUE' if val else 'FALSE')
                elif isinstance(val, (int, float)):
                    values.append(str(val))
                elif isinstance(val, datetime):
                    values.append(f"'{val.isoformat()}'")
                elif '🔒' in str(val):  # Encrypted field placeholder
                    values.append("'ENCRYPTED_PLACEHOLDER'")
                else:
                    # Escape single quotes
                    safe_val = str(val).replace("'", "''")
                    values.append(f"'{safe_val}'")
            
            sql = f"INSERT INTO {table_name} ({', '.join(column_names)}) VALUES ({', '.join(values)});"
            insert_statements.append(sql)
        
        return insert_statements
    
    def migrate_table(self, sheet_name: str):
        """Migrate a single table from Excel to PostgreSQL"""
        print(f"\n📊 Migrating table: {sheet_name}")
        
        # Extract definition
        table_def = self.extract_table_definition(sheet_name)
        table_name = table_def['table_name']
        
        print(f"  Table: {table_name}")
        print(f"  Columns: {len(table_def['columns'])}")
        print(f"  Sample rows: {len(table_def['sample_data'])}")
        
        # Generate SQL
        create_sql = self.generate_create_table_sql(table_def)
        insert_sqls = self.generate_insert_sample_data_sql(table_def)
        
        # Execute
        cursor = self.conn.cursor()
        
        try:
            # Create table
            print(f"  Creating table...")
            cursor.execute(create_sql)
            
            # Insert sample data
            if insert_sqls:
                print(f"  Inserting {len(insert_sqls)} sample rows...")
                for sql in insert_sqls:
                    cursor.execute(sql)
            
            self.conn.commit()
            print(f"  ✅ Successfully migrated {table_name}")
            
        except Exception as e:
            self.conn.rollback()
            print(f"  ❌ Error migrating {table_name}: {e}")
            raise
        
        finally:
            cursor.close()
    
    def migrate_all(self):
        """Migrate all tables from Excel to PostgreSQL"""
        print("\n" + "="*60)
        print("CHENU Excel to PostgreSQL Migration")
        print("="*60)
        
        self.connect()
        self.load_excel()
        
        # Table sheets to migrate (in order due to foreign keys)
        table_sheets = [
            '7_USERS',
            '2_LLM_PROVIDERS',
            '9_LLM_MODELS',
            '1_AGENTS',
            '3_AGENT_INTEGRATIONS',
            '6_WORKFLOWS',
            '5_TASKS',
            '4_AGENT_USAGE_LOGS',
            '8_BUDGET_ALERTS'
        ]
        
        success_count = 0
        failed_count = 0
        
        for sheet in table_sheets:
            try:
                self.migrate_table(sheet)
                success_count += 1
            except Exception as e:
                failed_count += 1
                print(f"Failed to migrate {sheet}")
        
        print("\n" + "="*60)
        print("Migration Summary")
        print("="*60)
        print(f"✅ Successful: {success_count}/{len(table_sheets)}")
        print(f"❌ Failed: {failed_count}/{len(table_sheets)}")
        
        if self.conn:
            self.conn.close()
            print("\nConnection closed.")
    
    def verify_migration(self):
        """Verify migration by checking row counts"""
        print("\n" + "="*60)
        print("Verification")
        print("="*60)
        
        self.connect()
        cursor = self.conn.cursor()
        
        tables = ['users', 'llm_providers', 'llm_models', 'agents', 
                  'agent_integrations', 'workflows', 'tasks', 
                  'agent_usage_logs', 'budget_alerts']
        
        for table in tables:
            cursor.execute(f"SELECT COUNT(*) FROM {table}")
            count = cursor.fetchone()[0]
            print(f"  {table}: {count} rows")
        
        cursor.close()
        self.conn.close()


# ═══════════════════════════════════════════════════════════════════════════
# USAGE EXAMPLE
# ═══════════════════════════════════════════════════════════════════════════

def main():
    """
    Main migration script
    
    Usage:
        python excel_to_postgresql_migration.py
    """
    
    # Configuration
    EXCEL_FILE = "/path/to/CHENU_DATABASE_DESIGN.xlsx"
    POSTGRES_CONNECTION = "postgresql://user:password@localhost:5432/chenu"
    
    # Create migrator
    migrator = ExcelToPostgreSQLMigration(EXCEL_FILE, POSTGRES_CONNECTION)
    
    # Run migration
    migrator.migrate_all()
    
    # Verify
    migrator.verify_migration()


# ═══════════════════════════════════════════════════════════════════════════
# CLI INTERFACE
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Migrate CHENU database from Excel to PostgreSQL')
    parser.add_argument('--excel', required=True, help='Path to Excel file')
    parser.add_argument('--db', required=True, help='PostgreSQL connection string')
    parser.add_argument('--verify-only', action='store_true', help='Only verify existing migration')
    
    args = parser.parse_args()
    
    migrator = ExcelToPostgreSQLMigration(args.excel, args.db)
    
    if args.verify_only:
        migrator.verify_migration()
    else:
        migrator.migrate_all()
        migrator.verify_migration()


"""
EXAMPLE USAGE:

# Basic usage
python excel_to_postgresql_migration.py \\
    --excel /path/to/CHENU_DATABASE_DESIGN.xlsx \\
    --db "postgresql://user:password@localhost:5432/chenu"

# Verify only (don't run migration)
python excel_to_postgresql_migration.py \\
    --excel /path/to/CHENU_DATABASE_DESIGN.xlsx \\
    --db "postgresql://user:password@localhost:5432/chenu" \\
    --verify-only

# With environment variables
export CHENU_DB="postgresql://user:password@localhost:5432/chenu"
python excel_to_postgresql_migration.py \\
    --excel CHENU_DATABASE_DESIGN.xlsx \\
    --db $CHENU_DB
"""

